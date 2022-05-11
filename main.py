from selenium import webdriver
import chromedriver_autoinstaller
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from time import sleep
import openpyxl


chromedriver_autoinstaller.install()
chromeWindow = webdriver.Chrome()

#Abrimos excel
bk = openpyxl.load_workbook("dummy.xlsx")
ex = bk.active


for i in range(1000):
    if i != 0 and i != 1:
        print(i)

        url = ex.cell(row=i, column=1)

        if url.value == None:
            print("empty cell")
            break
        else:
            print(url.value)
            ########################### Registrar RFC ##########################
            trueRfc = False
            RFC = ""
            chromeWindow.get(url.value)
            valueRfc = str(chromeWindow.find_element(By.XPATH, '/html/body/div[1]/div/form/ul[1]/li').text)
            for j in range(len(valueRfc)):
                if valueRfc[j-1] == ':':
                    trueRfc = True
                elif valueRfc[j] == ',':
                    trueRfc = False

                if trueRfc == True:
                    RFC += valueRfc[j]

            print(RFC)

            rfcCell = ex.cell(row=i, column=2)
            rfcCell.value = RFC

            ################# Registramos Curp #####################
            valueCurp = str(chromeWindow.find_element(By.XPATH, '/html/body/div[1]/div/form/ul[2]/li[2]/table/tbody/tr[1]/td/div/div/table/tbody/tr[1]/td[2]').text)
            print(valueCurp)
            curpCell = ex.cell(row=i, column=3)
            curpCell.value = valueCurp

            ################# Registramos Apellido Paterno ############
            valueApellidoPat = str(chromeWindow.find_element(By.XPATH, '/html/body/div[1]/div/form/ul[2]/li[2]/table/tbody/tr[1]/td/div/div/table/tbody/tr[3]/td[2]').text)
            print(valueApellidoPat)
            paternoCell = ex.cell(row=i, column=4)
            paternoCell.value = valueApellidoPat

            ################# Registramos Apellido Materno ############
            valueApellidoMat = str(chromeWindow.find_element(By.XPATH, '/html/body/div[1]/div/form/ul[2]/li[2]/table/tbody/tr[1]/td/div/div/table/tbody/tr[4]/td[2]').text)
            print(valueApellidoMat)
            maternoCell = ex.cell(row=i, column=5)
            maternoCell.value = valueApellidoMat

            ############## Registramos Nombres ####################
            valueNombres = str(chromeWindow.find_element(By.XPATH, '/html/body/div[1]/div/form/ul[2]/li[2]/table/tbody/tr[1]/td/div/div/table/tbody/tr[2]/td[2]').text)
            print(valueNombres)
            nombresCell = ex.cell(row=i, column=6)
            nombresCell.value = valueNombres

            ############## Registramos Codigo postal #############
            valueCP = str(chromeWindow.find_element(By.XPATH, '/html/body/div[1]/div/form/ul[3]/li[2]/table/tbody/tr[1]/td/div/div/table/tbody/tr[8]/td[2]').text)
            print(valueCP)
            cpCell = ex.cell(row=i, column=7)
            cpCell.value = valueCP

            ################## Registramos Calle ####################
            valueCalle = str(chromeWindow.find_element(By.XPATH, '/html/body/div[1]/div/form/ul[3]/li[2]/table/tbody/tr[1]/td/div/div/table/tbody/tr[5]/td[2]').text)
            print(valueCalle)
            calleCell = ex.cell(row=i, column=8)
            calleCell.value = valueCalle

            #################### Registramos Numero exterior ###################
            valueNumExt = str(chromeWindow.find_element(By.XPATH, '/html/body/div[1]/div/form/ul[3]/li[2]/table/tbody/tr[1]/td/div/div/table/tbody/tr[6]/td[2]').text)
            print(valueNumExt)
            numExt = ex.cell(row=i, column=9)
            numExt.value = valueNumExt


            ##################### Registramos Colonia ########################
            valueCol = str(chromeWindow.find_element(By.XPATH, '/html/body/div[1]/div/form/ul[3]/li[2]/table/tbody/tr[1]/td/div/div/table/tbody/tr[3]/td[2]').text)
            print(valueCol)
            colCell = ex.cell(row=i,column=10)
            colCell.value = valueCol

            ################### Registramos Municipio #####################
            valueMun = str(chromeWindow.find_element(By.XPATH, '/html/body/div[1]/div/form/ul[3]/li[2]/table/tbody/tr[1]/td/div/div/table/tbody/tr[2]/td[2]').text)
            print(valueMun)
            munCell = ex.cell(row=i, column=11)
            munCell.value = valueMun

            ################### Registramos Entidad federativa ################
            valueEntidad = str(chromeWindow.find_element(By.XPATH, '/html/body/div[1]/div/form/ul[3]/li[2]/table/tbody/tr[1]/td/div/div/table/tbody/tr[1]/td[2]').text)
            print(valueEntidad)
            entidadCell = ex.cell(row=i, column=12)
            entidadCell.value = valueEntidad


            #################### Regimen fiscal #########################
            valueRegimen = str( chromeWindow.find_element(By.XPATH, '/html/body/div[1]/div/form/ul[4]/li[2]/table/tbody/tr[1]/td/div/div/table/tbody/tr[1]/td[2]').text)
            print(valueRegimen)
            regimenCell = ex.cell(row=i,column=13)
            regimenCell.value = valueRegimen






            bk.save('Situacion Fiscal.xlsx')








